from openpyxl import load_workbook

months = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Agu", "Sep", "Oct", "Nov", "Dec"]
wk_to_take_values = None
wk_to_receive_values = None

def load_work_books():
    global wk_to_take_values
    global wk_to_receive_values

    wk_to_take_values = load_workbook(filename = 'planilha-rentabilidade-2022-Editavel.xlsm', data_only=True)
    wk_to_receive_values = load_workbook(filename = "planilha-rentabilidade-2022-Consolidado-Editavel.xlsm")

def run_job():
    row_init_reference: int = 3
    rows_in_each_month: int = 4
    rows_to_next_month: int = 4

    for month in months:
        for row in range(rows_in_each_month):
            cell_row = row + 1
            print("Mes {} linha {}".format(month, cell_row))

            cell_to_get = "S{}".format(row_init_reference)
            cell_to_set = "C{}".format(row_init_reference)
            print("Get from Cell {} to set in Cell {}".format(cell_to_get, cell_to_set))
            # chamar a função para pegar o valor e jogar na outra planilha
            value_to_send = get_value_from_wk(cell_to_get)
            set_value_to_receive_wk(cell_to_set, value_to_send)
            row_init_reference = row_init_reference  + 1

        row_init_reference = row_init_reference + rows_to_next_month

def get_value_from_wk(cell):
    aba_active_to_take = wk_to_take_values.active
    print(f"Lendo a tabela Para Pegar Dados na aba {aba_active_to_take.title}")
    value_took = aba_active_to_take[cell].value
    print(f"Pegando o valor a ser transferido = {value_took}")
    return value_took

def set_value_to_receive_wk(cell, value):
    if (value != None):
        aba_active_to_receive = wk_to_receive_values.active
        print(f"Abrindo a tabela Consolidado na aba {aba_active_to_receive.title} para receber o valor")
        print(f"Salvando o valor na nova tabela = {value}")
        aba_active_to_receive[cell] = value

def save_to_new_wk():
    wk_to_receive_values.save("planilha-rentabilidade-2022-Consolidado-Editavel-copy.xlsx")

def init():
    load_work_books()
    run_job()
    save_to_new_wk()

init()
