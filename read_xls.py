from openpyxl import load_workbook

wb_1 = load_workbook(filename = 'planilha-rentabilidade-2022-Editavel.xlsm', data_only=True)

aba_active_1 = wb_1.active

print(f"Lendo a tabela na aba {aba_active_1.title}")
total_value_1 = aba_active_1["S30"].value
print(f"Pegando o valor a ser transferido = {total_value_1}")

wb_2 = load_workbook(filename = "planilha-rentabilidade-2022-Consolidado-Editavel.xlsm")

aba_active_2 = wb_2.active
print(f"Abrindo a tabela Consolidado na aba {aba_active_2.title} para receber o valor")

aba_active_2["C30"] = total_value_1

print(aba_active_2["C30"].value)

wb_2.save("planilha-rentabilidade-2022-Consolidado-Editavel-3.xlsx")
# for cell in aba_active["S"]:
#     print(cell)